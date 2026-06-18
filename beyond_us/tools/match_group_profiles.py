# 조별 명단과 Supabase 앱 가입자를 이름과 교구 기준으로 대조한다.
from __future__ import annotations

import json
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
GROUP_FILE = Path(r"C:\Users\jkjk9\Downloads\조별 데이터 정리 6_14.xlsx")
PROFILES_FILE = ROOT / "beyond_us" / "tmp_profiles_prod.json"
OUTPUT_DIR = ROOT / "beyond_us" / "outputs"
OUTPUT_FILE = OUTPUT_DIR / "group_profile_matching_20260618.xlsx"
SUMMARY_FILE = OUTPUT_DIR / "group_profile_matching_20260618_summary.json"


@dataclass
class RosterPerson:
    source_sheet: str
    group_no: str
    role: str
    name: str
    birth_year: str
    parish_raw: str
    parish_norm: str
    phone: str
    schedule: str
    note: str
    source_row: int


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_name(value: str) -> str:
    return re.sub(r"\s+", "", clean(value)).lower()


def normalize_parish(value: str) -> str:
    text = clean(value)
    compact = re.sub(r"\s+", "", text)
    if not compact:
        return ""
    upper = compact.upper()
    if "VIP" in upper:
        return "VIP"
    if "교회학교" in compact:
        return "교회학교"
    if "목양" in compact:
        return "목양교구"
    match = re.search(r"([1-4])(?:청|교구|청년)", compact)
    if match:
        return f"{match.group(1)}청"
    return text


def group_sort_key(group_no: str) -> tuple[int, str]:
    match = re.search(r"\d+", group_no)
    if match:
        return (int(match.group()), group_no)
    return (999, group_no)


def profile_label(profile: dict[str, Any]) -> str:
    parts = [
        f"{profile.get('login_id', '')}",
        f"{profile.get('name', '')}",
        f"{profile.get('parish', '')}",
        f"{profile.get('participant_code', '')}",
    ]
    return " / ".join(clean(v) for v in parts if clean(v))


def extract_group_roster(path: Path) -> list[RosterPerson]:
    wb = openpyxl.load_workbook(path, data_only=True)
    roster: list[RosterPerson] = []

    for sheet_name in ["1~8조", "9~16조"]:
        ws = wb[sheet_name]
        labels: list[tuple[int, int, str]] = []
        for row in ws.iter_rows():
            for cell in row:
                value = clean(cell.value)
                if re.fullmatch(r"\d+조", value):
                    labels.append((cell.row, cell.column, value))

        for label_row, label_col, group_no in labels:
            next_rows = [
                row
                for row, col, _ in labels
                if col == label_col and row > label_row
            ]
            end_row = min(next_rows) - 1 if next_rows else ws.max_row
            for row_idx in range(label_row + 2, end_row + 1):
                name = clean(ws.cell(row_idx, label_col + 1).value)
                if not name:
                    continue
                role = clean(ws.cell(row_idx, label_col).value)
                parish_raw = clean(ws.cell(row_idx, label_col + 3).value)
                roster.append(
                    RosterPerson(
                        source_sheet=sheet_name,
                        group_no=group_no,
                        role=role,
                        name=name,
                        birth_year=clean(ws.cell(row_idx, label_col + 2).value),
                        parish_raw=parish_raw,
                        parish_norm=normalize_parish(parish_raw),
                        phone=clean(ws.cell(row_idx, label_col + 4).value),
                        schedule=clean(ws.cell(row_idx, label_col + 5).value),
                        note=clean(ws.cell(row_idx, label_col + 6).value),
                        source_row=row_idx,
                    )
                )

    if "2차설문 추가 필요 명단" in wb.sheetnames:
        ws = wb["2차설문 추가 필요 명단"]
        category = ""
        for row_idx in range(8, ws.max_row + 1):
            if clean(ws.cell(row_idx, 2).value):
                category = clean(ws.cell(row_idx, 2).value)
            name = clean(ws.cell(row_idx, 4).value)
            if not name:
                continue
            parish_raw = clean(ws.cell(row_idx, 6).value)
            roster.append(
                RosterPerson(
                    source_sheet=ws.title,
                    group_no="추가필요",
                    role=category,
                    name=name,
                    birth_year=clean(ws.cell(row_idx, 5).value),
                    parish_raw=parish_raw,
                    parish_norm=normalize_parish(parish_raw),
                    phone=clean(ws.cell(row_idx, 7).value),
                    schedule=clean(ws.cell(row_idx, 8).value),
                    note=clean(ws.cell(row_idx, 9).value),
                    source_row=row_idx,
                )
            )

    return roster


def load_profiles(path: Path) -> list[dict[str, Any]]:
    profiles = json.loads(path.read_text(encoding="utf-8-sig"))
    filtered = []
    for profile in profiles:
        if profile.get("account_status") != "active":
            continue
        if profile.get("is_dev") or profile.get("is_test"):
            continue
        p = dict(profile)
        p["name_norm"] = normalize_name(p.get("name", ""))
        p["parish_norm"] = normalize_parish(p.get("parish", ""))
        filtered.append(p)
    return filtered


def build_report(roster: list[RosterPerson], profiles: list[dict[str, Any]]) -> dict[str, Any]:
    profiles_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for profile in profiles:
        profiles_by_name[profile["name_norm"]].append(profile)

    roster_by_name: dict[str, list[RosterPerson]] = defaultdict(list)
    for person in roster:
        roster_by_name[normalize_name(person.name)].append(person)

    matched_profile_ids: set[str] = set()
    rows: list[dict[str, Any]] = []
    missing_nickname: list[dict[str, Any]] = []
    duplicate_needs_check: list[dict[str, Any]] = []

    for person in sorted(roster, key=lambda p: (group_sort_key(p.group_no), p.role, p.name, p.source_row)):
        name_norm = normalize_name(person.name)
        candidates = profiles_by_name.get(name_norm, [])
        parish_candidates = [p for p in candidates if p["parish_norm"] == person.parish_norm]
        status = ""
        selected: dict[str, Any] | None = None
        candidate_text = "; ".join(profile_label(p) for p in candidates)

        if not candidates:
            status = "닉네임 없음"
        elif len(candidates) == 1:
            selected = candidates[0]
            status = "매칭"
        elif len(parish_candidates) == 1:
            selected = parish_candidates[0]
            status = "교구 기준 매칭"
        elif len(parish_candidates) > 1:
            status = "이름 중복 확인필요 - 같은 청"
        else:
            status = "이름 중복 확인필요 - 다른 청 후보"

        if selected:
            matched_profile_ids.add(selected["id"])

        row = {
            "조": person.group_no,
            "역할": person.role,
            "이름": person.name,
            "년생": person.birth_year,
            "조_교구": person.parish_raw,
            "정규화_청": person.parish_norm,
            "참여일정": person.schedule,
            "특이사항": person.note,
            "매칭상태": status,
            "닉네임": selected.get("login_id", "") if selected else "",
            "앱_이름": selected.get("name", "") if selected else "",
            "앱_교구": selected.get("parish", "") if selected else "",
            "앱_번호": selected.get("participant_code", "") if selected else "",
            "후보": candidate_text,
            "원본시트": person.source_sheet,
            "원본행": person.source_row,
        }
        rows.append(row)
        if status == "닉네임 없음":
            missing_nickname.append(row)
        if status.startswith("이름 중복"):
            duplicate_needs_check.append(row)

    roster_names = set(roster_by_name)
    non_attendees = []
    for profile in sorted(profiles, key=lambda p: (p["parish_norm"], p.get("name", ""), p.get("login_id", ""))):
        if profile["id"] in matched_profile_ids:
            continue
        if profile["name_norm"] in roster_names:
            continue
        non_attendees.append(
            {
                "상태": "미참",
                "닉네임": profile.get("login_id", ""),
                "이름": profile.get("name", ""),
                "교구": profile.get("parish", ""),
                "정규화_청": profile.get("parish_norm", ""),
                "권한": profile.get("role", ""),
                "앱_번호": profile.get("participant_code", ""),
                "가입일": profile.get("created_at", ""),
            }
        )

    app_duplicate_rows = []
    for name_norm, people in sorted(profiles_by_name.items(), key=lambda kv: kv[0]):
        if len(people) <= 1:
            continue
        parish_counts: dict[str, int] = defaultdict(int)
        for person in people:
            parish_counts[person["parish_norm"]] += 1
        for person in people:
            same_parish = parish_counts[person["parish_norm"]] > 1
            app_duplicate_rows.append(
                {
                    "구분": "앱가입자 이름 중복",
                    "이름": person.get("name", ""),
                    "닉네임": person.get("login_id", ""),
                    "교구": person.get("parish", ""),
                    "정규화_청": person.get("parish_norm", ""),
                    "앱_번호": person.get("participant_code", ""),
                    "표기": "청까지 같음" if same_parish else "청으로 구분 가능",
                }
            )

    roster_duplicate_rows = []
    for name_norm, people in sorted(roster_by_name.items(), key=lambda kv: kv[0]):
        if len(people) <= 1:
            continue
        parish_counts: dict[str, int] = defaultdict(int)
        for person in people:
            parish_counts[person.parish_norm] += 1
        for person in people:
            same_parish = parish_counts[person.parish_norm] > 1
            roster_duplicate_rows.append(
                {
                    "구분": "조명단 이름 중복",
                    "이름": person.name,
                    "조": person.group_no,
                    "역할": person.role,
                    "교구": person.parish_raw,
                    "정규화_청": person.parish_norm,
                    "표기": "청까지 같음" if same_parish else "청으로 구분 가능",
                }
            )

    return {
        "matched_rows": rows,
        "missing_nickname": missing_nickname,
        "non_attendees": non_attendees,
        "duplicate_needs_check": duplicate_needs_check,
        "app_name_duplicates": app_duplicate_rows,
        "roster_name_duplicates": roster_duplicate_rows,
        "summary": {
            "roster_count": len(roster),
            "active_profile_count": len(profiles),
            "matched_count": len([r for r in rows if r["매칭상태"] in ("매칭", "교구 기준 매칭")]),
            "missing_nickname_count": len(missing_nickname),
            "duplicate_needs_check_count": len(duplicate_needs_check),
            "non_attendee_count": len(non_attendees),
            "app_name_duplicate_count": len(app_duplicate_rows),
            "roster_name_duplicate_count": len(roster_duplicate_rows),
        },
    }


def write_sheet(ws, rows: list[dict[str, Any]], headers: list[str], title_fill: str = "1F4E79") -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor=title_fill)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            max_len = max(max_len, len(clean(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 42)

    for row in ws.iter_rows(min_row=2):
        status = clean(row[headers.index("매칭상태")].value) if "매칭상태" in headers else ""
        if status == "닉네임 없음":
            fill = PatternFill("solid", fgColor="FFF2CC")
        elif status.startswith("이름 중복"):
            fill = PatternFill("solid", fgColor="F4CCCC")
        else:
            fill = None
        if fill:
            for cell in row:
                cell.fill = fill


def write_workbook(report: dict[str, Any]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "요약"
    summary_rows = [{"항목": key, "값": value} for key, value in report["summary"].items()]
    write_sheet(summary_ws, summary_rows, ["항목", "값"], "5B21B6")

    write_sheet(
        wb.create_sheet("조별 매칭 결과"),
        report["matched_rows"],
        ["조", "역할", "이름", "년생", "조_교구", "정규화_청", "참여일정", "특이사항", "매칭상태", "닉네임", "앱_이름", "앱_교구", "앱_번호", "후보", "원본시트", "원본행"],
    )
    write_sheet(
        wb.create_sheet("닉네임 없음"),
        report["missing_nickname"],
        ["조", "역할", "이름", "년생", "조_교구", "정규화_청", "참여일정", "특이사항", "매칭상태", "후보", "원본시트", "원본행"],
        "B45309",
    )
    write_sheet(
        wb.create_sheet("미참"),
        report["non_attendees"],
        ["상태", "닉네임", "이름", "교구", "정규화_청", "권한", "앱_번호", "가입일"],
        "7F1D1D",
    )
    write_sheet(
        wb.create_sheet("이름 중복 확인필요"),
        report["duplicate_needs_check"],
        ["조", "역할", "이름", "년생", "조_교구", "정규화_청", "참여일정", "특이사항", "매칭상태", "후보", "원본시트", "원본행"],
        "991B1B",
    )
    write_sheet(
        wb.create_sheet("앱 이름 중복"),
        report["app_name_duplicates"],
        ["구분", "이름", "닉네임", "교구", "정규화_청", "앱_번호", "표기"],
        "92400E",
    )
    write_sheet(
        wb.create_sheet("조명단 이름 중복"),
        report["roster_name_duplicates"],
        ["구분", "이름", "조", "역할", "교구", "정규화_청", "표기"],
        "92400E",
    )
    wb.save(OUTPUT_FILE)
    SUMMARY_FILE.write_text(json.dumps(report["summary"], ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    roster = extract_group_roster(GROUP_FILE)
    profiles = load_profiles(PROFILES_FILE)
    report = build_report(roster, profiles)
    write_workbook(report)
    print(json.dumps({"output": str(OUTPUT_FILE), **report["summary"]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
